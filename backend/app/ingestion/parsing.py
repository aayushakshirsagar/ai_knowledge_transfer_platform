## after parsing - a document object is returned - not flattening it to plain text 
from __future__ import annotations

from pathlib import Path
from typing import Any, Literal

import pymupdf
from pydantic import BaseModel, Field

from docling.datamodel.base_models import InputFormat
from docling.datamodel.pipeline_options import PdfPipelineOptions
import os
os.environ["TORCHDYNAMO_DISABLE"] = "1"

from docling.document_converter import (
    DocumentConverter, 
    PdfFormatOption,
)

import csv
from openpyxl import load_workbook


SectionType = Literal["heading", "paragraph", "table"]

class DocumentSection(BaseModel):

    type: SectionType
    content: Any
    metadata: dict[str, Any] = Field(default_factory=dict)


class ParsedDocument(BaseModel):
    #object that will go to chunking
    sections: list[DocumentSection] = Field(default_factory=list)

def _build_docling_converter() -> DocumentConverter:
    # primary Docling converter.
     """
     - preserve table structure 
     - OCR is currently disabled 
     - Scanned doc ocr will be added separately 

     """

     pdf_pipeline_options = PdfPipelineOptions()

     # Preserving the table structure
     pdf_pipeline_options.do_table_structure = True

     #Enabling OCR for scanned/image-only PDFs
     pdf_pipeline_options.do_ocr = False
     #pdf_pipeline_options.ocr_options = TesseractOcrOptions()

     return DocumentConverter(
        format_options={
            InputFormat.PDF: PdfFormatOption(
                pipeline_options=pdf_pipeline_options,
            
         ),
        }
     )

def _parse_with_docling(file_path: Path) -> ParsedDocument:
    """
    Parse a document using Docling and return a ParsedDocument object.
    """
    converter = _build_docling_converter()

    result = converter.convert(str(file_path))

    

    return _convert_docling_document(result.document, file_path)

def _convert_docling_document(docling_document, file_path: Path,)-> ParsedDocument:
    """
    Convert a Docling document to a ParsedDocument object.
    """
    sections: list[DocumentSection] = []
    source_type = file_path.suffix.lower().lstrip(".")

    for item, _level in docling_document.iterate_items():
        label = getattr(item, "label", None)
        text = getattr(item, "text", None)

        # Build metadata
        # ---------------------------------------------------------

        metadata: dict[str, Any] = {
            "source_type": source_type,
            "file_name": file_path.name,
        }

        # Docling items can contain provenance information.
        provenance = getattr(item, "prov", None)

        if provenance:
            try:
                # Usually the first provenance entry contains
                # the relevant page information.
                first_provenance = provenance[0]

                page_no = getattr(
                    first_provenance,
                    "page_no",
                    None,
                )

                if page_no is not None:
                    metadata["page"] = page_no

            except (IndexError, TypeError):
                pass

        if label in ("title", "section_header"):
            if text and text.strip():
                sections.append(
                    DocumentSection(
                        type="heading",
                        content=text.strip(),
                        metadata=metadata,
                    )
                )

        elif label == "table":
                table = _convert_docling_table(item)

                if table:
                    sections.append(
                        DocumentSection(
                            type="table",
                            content=table,
                            metadata=metadata,
                        )
                    )

        else:
                if text and text.strip():
                    sections.append(
                        DocumentSection(
                            type="paragraph",
                            content=text.strip(),
                            metadata=metadata,
                        )
                    )

       

    return ParsedDocument(sections=sections)

def _parse_with_csv(file_path: Path) -> ParsedDocument:
    """
    Parse a CSV file as a structured table.
    """

    sections: list[DocumentSection] = []

    with file_path.open(
        "r",
        encoding="utf-8-sig",
        newline="",
    ) as file:

        reader = csv.reader(file)

        rows = [
            [str(cell) for cell in row]
            for row in reader
        ]

    if rows:
        sections.append(
            DocumentSection(
                type="table",
                content=rows,
                metadata={
                    "source_type": "csv",
                },
            )
        )

    return ParsedDocument(sections=sections)

def _parse_with_xlsx(file_path: Path) -> ParsedDocument:
    """
    Parse an Excel workbook.

    Each worksheet becomes a table section.
    """

    sections: list[DocumentSection] = []

    workbook = load_workbook(
        filename=file_path,
        read_only=True,
        data_only=True,
    )

    for worksheet in workbook.worksheets:

        rows: list[list[str]] = []

        for row in worksheet.iter_rows(values_only=True):

            values = [
                "" if value is None else str(value)
                for value in row
            ]

            # Ignore completely empty rows.
            if any(value.strip() for value in values):
                rows.append(values)

        if rows:
            sections.append(
                DocumentSection(
                    type="table",
                    content=rows,
                    metadata={
                        "source_type": "xlsx",
                        "sheet_name": worksheet.title,
                    },
                )
            )

    workbook.close()

    return ParsedDocument(sections=sections)

def _parse_with_text(file_path: Path) -> ParsedDocument:
    """
    Parse plain text or Markdown.
    """

    text = file_path.read_text(
        encoding="utf-8"
    ).strip()

    if not text:
        return ParsedDocument()

    return ParsedDocument(
        sections=[
            DocumentSection(
                type="paragraph",
                content=text,
                metadata={
                    "source_type": file_path.suffix.lower().lstrip("."),
                },
            )
        ]
    )



def _convert_docling_table(table_item)-> list[list[str]]:
    """
    Convert a Docling table item to a list of lists (rows and columns).
    """
    try:
        dataframe = table_item.export_to_dataframe()

        dataframe = dataframe.fillna("")

        return [
            [str(value) for value in row]
            for row in dataframe.values.tolist()

            
            
        ]
    except Exception:
        # for a table that cannot be converted 

        return []

#PyMuPdf parsing fallback

def _parse_with_pymupdf(file_path: Path) -> ParsedDocument:
    sections: list[DocumentSection] = []

    with pymupdf.open(file_path) as document:

        for page_number, page in enumerate(document, start=1):

            text = page.get_text("text").strip()

            if not text:
                continue

            sections.append(
                DocumentSection(
                    type="paragraph",
                    content=text,
                    metadata={
                        "source_type":"pdf",
                        "file_name": file_path.name,
                        "page": page_number,
                    }
                )
            )

    return ParsedDocument(sections=sections)

## Parsing API 
def parse_document(
    file_path: str,
    mime_type: str,
) -> ParsedDocument:

    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(
            f"Document does not exist: {file_path}"
        )

    if not path.is_file():
        raise ValueError(
            f"Document path is not a file: {file_path}"
        )

    suffix = path.suffix.lower()

    # ---------------------------------------------------------
    # Docling formats
    # ---------------------------------------------------------

    if suffix in {".pdf", ".docx", ".pptx"}:

        try:
            parsed_document = _parse_with_docling(path)

            if parsed_document.sections:
                return parsed_document

        except Exception as docling_error:

            # PDF has PyMuPDF fallback.
            if suffix == ".pdf":

                try:
                    parsed_document = _parse_with_pymupdf(path)

                    if parsed_document.sections:
                        return parsed_document

                except Exception as pymupdf_error:

                    raise RuntimeError(
                        "Docling failed and PyMuPDF fallback also failed."
                    ) from pymupdf_error

            raise RuntimeError(
                f"Unable to parse document with Docling: {file_path}"
            ) from docling_error

    # ---------------------------------------------------------
    # CSV
    # ---------------------------------------------------------

    if suffix == ".csv":

        return _parse_with_csv(path)

    # ---------------------------------------------------------
    # Excel
    # ---------------------------------------------------------

    if suffix in {".xlsx", ".xlsm"}:

        return _parse_with_xlsx(path)

    # ---------------------------------------------------------
    # Plain text / Markdown
    # ---------------------------------------------------------

    if suffix in {".txt", ".md"}:

        return _parse_with_text(path)

    # ---------------------------------------------------------
    # Unsupported
    # ---------------------------------------------------------

    raise ValueError(
        f"Unsupported document format: {suffix}"
    )